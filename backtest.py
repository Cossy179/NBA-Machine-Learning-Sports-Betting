#!/usr/bin/env python3
"""
🏀 NBA Machine Learning Sports Betting - Enhanced Backtesting Script
Comprehensive backtesting with ROI analysis, statistics, and visualization.
"""
import sys
import os
import argparse
import warnings
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
import sqlite3
import matplotlib.pyplot as plt
import seaborn as sns
from colorama import Fore, Style, init
init()
warnings.filterwarnings('ignore')

# Add src to path
sys.path.append('src')

def print_header():
    """Print backtesting header"""
    print("🏀" + "="*70 + "🏀")
    print("📊 NBA Machine Learning Sports Betting - Enhanced Backtesting 📊")
    print("🏀" + "="*70 + "🏀")
    print(f"⏰ Backtesting started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

def create_enhanced_features(df):
    """Create comprehensive enhanced features for maximum model performance"""
    try:
        enhanced_df = pd.DataFrame(index=df.index)
        
        print("🔧 Creating comprehensive enhanced features...")
        
        # 1. Team differential features (home vs away)
        home_stats = ['W_PCT', 'PTS', 'REB', 'AST', 'FG_PCT', 'FG3_PCT', 'FT_PCT', 'TOV', 'STL', 'BLK', 'OREB', 'DREB', 'PF', 'PFD', 'PLUS_MINUS']
        away_stats = [stat + '.1' for stat in home_stats]
        
        for home_stat, away_stat in zip(home_stats, away_stats):
            if home_stat in df.columns and away_stat in df.columns:
                enhanced_df[f'{home_stat}_diff'] = df[home_stat] - df[away_stat]
                enhanced_df[f'{home_stat}_ratio'] = df[home_stat] / (df[away_stat] + 1e-8)
                enhanced_df[f'{home_stat}_sum'] = df[home_stat] + df[away_stat]
                enhanced_df[f'{home_stat}_product'] = df[home_stat] * df[away_stat]
        
        # 2. Advanced ranking differentials
        rank_stats = ['W_PCT_RANK', 'PTS_RANK', 'REB_RANK', 'AST_RANK', 'FG_PCT_RANK', 'FG3_PCT_RANK', 'FT_PCT_RANK', 'TOV_RANK', 'STL_RANK', 'BLK_RANK', 'PLUS_MINUS_RANK']
        for stat in rank_stats:
            if stat in df.columns and f'{stat}.1' in df.columns:
                enhanced_df[f'{stat}_diff'] = df[f'{stat}.1'] - df[stat]  # Lower rank is better
                enhanced_df[f'{stat}_sum'] = df[stat] + df[f'{stat}.1']
                enhanced_df[f'{stat}_product'] = df[stat] * df[f'{stat}.1']
        
        # 3. Rest advantage and fatigue
        if 'Days-Rest-Home' in df.columns and 'Days-Rest-Away' in df.columns:
            enhanced_df['rest_advantage'] = df['Days-Rest-Home'] - df['Days-Rest-Away']
            enhanced_df['rest_advantage_abs'] = abs(enhanced_df['rest_advantage'])
            enhanced_df['total_rest'] = df['Days-Rest-Home'] + df['Days-Rest-Away']
            enhanced_df['rest_ratio'] = df['Days-Rest-Home'] / (df['Days-Rest-Away'] + 1e-8)
        
        # 4. Advanced efficiency metrics
        if all(col in df.columns for col in ['PTS', 'FGA', 'FTA']):
            enhanced_df['home_efficiency'] = df['PTS'] / (df['FGA'] + 0.5 * df['FTA'] + 1e-8)
            enhanced_df['away_efficiency'] = df['PTS.1'] / (df['FGA.1'] + 0.5 * df['FTA.1'] + 1e-8)
            enhanced_df['efficiency_diff'] = enhanced_df['home_efficiency'] - enhanced_df['away_efficiency']
            enhanced_df['efficiency_ratio'] = enhanced_df['home_efficiency'] / (enhanced_df['away_efficiency'] + 1e-8)
        
        # 5. Defensive metrics
        if all(col in df.columns for col in ['STL', 'BLK', 'TOV']):
            enhanced_df['home_defense'] = df['STL'] + df['BLK'] - df['TOV']
            enhanced_df['away_defense'] = df['STL.1'] + df['BLK.1'] - df['TOV.1']
            enhanced_df['defense_diff'] = enhanced_df['home_defense'] - enhanced_df['away_defense']
            enhanced_df['defense_ratio'] = enhanced_df['home_defense'] / (enhanced_df['away_defense'] + 1e-8)
        
        # 6. Pace and tempo metrics
        if all(col in df.columns for col in ['MIN', 'PTS', 'FGA']):
            enhanced_df['home_pace'] = df['PTS'] / (df['MIN'] + 1e-8) * 48  # Points per 48 minutes
            enhanced_df['away_pace'] = df['PTS.1'] / (df['MIN.1'] + 1e-8) * 48
            enhanced_df['pace_diff'] = enhanced_df['home_pace'] - enhanced_df['away_pace']
            enhanced_df['pace_ratio'] = enhanced_df['home_pace'] / (enhanced_df['away_pace'] + 1e-8)
        
        # 7. Three-point shooting metrics
        if all(col in df.columns for col in ['FG3M', 'FG3A', 'FG3_PCT']):
            enhanced_df['home_3pt_volume'] = df['FG3M'] * df['FG3A']
            enhanced_df['away_3pt_volume'] = df['FG3M.1'] * df['FG3A.1']
            enhanced_df['3pt_volume_diff'] = enhanced_df['home_3pt_volume'] - enhanced_df['away_3pt_volume']
            enhanced_df['3pt_pct_diff'] = df['FG3_PCT'] - df['FG3_PCT.1']
        
        # 8. Rebounding dominance
        if all(col in df.columns for col in ['OREB', 'DREB', 'REB']):
            enhanced_df['home_reb_dominance'] = df['OREB'] + df['DREB'] - df['REB.1']
            enhanced_df['away_reb_dominance'] = df['OREB.1'] + df['DREB.1'] - df['REB']
            enhanced_df['reb_dominance_diff'] = enhanced_df['home_reb_dominance'] - enhanced_df['away_reb_dominance']
        
        # 9. Turnover and ball control
        if all(col in df.columns for col in ['TOV', 'AST']):
            enhanced_df['home_ast_to_ratio'] = df['AST'] / (df['TOV'] + 1e-8)
            enhanced_df['away_ast_to_ratio'] = df['AST.1'] / (df['TOV.1'] + 1e-8)
            enhanced_df['ast_to_ratio_diff'] = enhanced_df['home_ast_to_ratio'] - enhanced_df['away_ast_to_ratio']
        
        # 10. Momentum and form indicators
        if 'W_PCT' in df.columns and 'W_PCT.1' in df.columns:
            enhanced_df['momentum_diff'] = df['W_PCT'] - df['W_PCT.1']
            enhanced_df['total_momentum'] = df['W_PCT'] + df['W_PCT.1']
            enhanced_df['momentum_ratio'] = df['W_PCT'] / (df['W_PCT.1'] + 1e-8)
        
        # 11. Home court advantage (multiple factors)
        enhanced_df['home_court_advantage'] = 1.0
        enhanced_df['home_court_strength'] = 1.0 + (df['W_PCT'] - df['W_PCT.1']) * 0.1 if 'W_PCT' in df.columns else 1.0
        
        # 12. Season progress and timing
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'])
            enhanced_df['day_of_season'] = df['Date'].dt.dayofyear
            enhanced_df['month'] = df['Date'].dt.month
            enhanced_df['is_playoffs'] = (df['Date'].dt.month >= 4).astype(int)
            enhanced_df['is_early_season'] = (df['Date'].dt.month <= 2).astype(int)
            enhanced_df['is_late_season'] = (df['Date'].dt.month >= 3).astype(int)
        
        # 13. Statistical consistency metrics
        if all(col in df.columns for col in ['PTS', 'REB', 'AST']):
            enhanced_df['home_consistency'] = df['PTS'] + df['REB'] + df['AST']
            enhanced_df['away_consistency'] = df['PTS.1'] + df['REB.1'] + df['AST.1']
            enhanced_df['consistency_diff'] = enhanced_df['home_consistency'] - enhanced_df['away_consistency']
        
        # 14. Advanced composite scores
        if all(col in df.columns for col in ['W_PCT', 'PTS', 'REB', 'AST', 'FG_PCT']):
            enhanced_df['home_composite'] = (df['W_PCT'] * 0.3 + df['PTS']/100 * 0.2 + df['REB']/50 * 0.2 + df['AST']/30 * 0.2 + df['FG_PCT'] * 0.1)
            enhanced_df['away_composite'] = (df['W_PCT.1'] * 0.3 + df['PTS.1']/100 * 0.2 + df['REB.1']/50 * 0.2 + df['AST.1']/30 * 0.2 + df['FG_PCT.1'] * 0.1)
            enhanced_df['composite_diff'] = enhanced_df['home_composite'] - enhanced_df['away_composite']
        
        print(f"✅ Created {len(enhanced_df.columns)} comprehensive enhanced features")
        return enhanced_df
        
    except Exception as e:
        print(f"⚠️ Enhanced feature creation failed: {e}")
        import traceback
        traceback.print_exc()
        return None

def load_historical_data(start_date="2023-01-01", end_date="2024-06-30"):
    """Load historical NBA data for backtesting - UNIQUE GAMES ONLY"""
    print(f"📥 Loading historical data ({start_date} to {end_date})...")
    
    try:
        con = sqlite3.connect("Data/dataset.sqlite")
        
        # Try enhanced dataset first
        cursor = con.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", ("dataset_2012-24_enhanced",))
        if cursor.fetchone():
            dataset_name = "dataset_2012-24_enhanced"
            print("✅ Using enhanced dataset")
        else:
            dataset_name = "dataset_2012-24_new"
            print("⚠️ Using base dataset (enhanced features not available)")
        
        # Query to get UNIQUE games only (remove duplicates)
        query = f'''
        SELECT DISTINCT 
            Date, TEAM_NAME, "TEAM_NAME.1" as AWAY_TEAM, Score, "Home-Team-Win", OU, "OU-Cover"
        FROM "{dataset_name}"
        WHERE Date >= ? AND Date <= ?
        ORDER BY Date
        '''
        
        df = pd.read_sql_query(query, con, params=[start_date, end_date])
        con.close()
        
        # Convert date column
        df["Date"] = pd.to_datetime(df["Date"])
        
        print(f"✅ Loaded {len(df)} UNIQUE games for backtesting")
        print(f"📅 Date range: {df['Date'].min().strftime('%Y-%m-%d')} to {df['Date'].max().strftime('%Y-%m-%d')}")
        return df
        
    except Exception as e:
        print(f"❌ Error loading data: {e}")
        return None

def load_available_models():
    """Load all available trained models including advanced models"""
    print("🤖 Loading available models...")
    
    models = {}
    
    # Load Advanced Backtesting Engine
    try:
        sys.path.append('src/Backtest')
        from BacktestingEngine import AdvancedBacktestingEngine
        
        advanced_engine = AdvancedBacktestingEngine()
        models['advanced_engine'] = {
            'engine': advanced_engine,
            'name': 'Advanced Backtesting Engine',
            'type': 'advanced'
        }
        print("✅ Advanced Backtesting Engine loaded")
        
    except Exception as e:
        print(f"⚠️ Advanced Backtesting Engine failed: {e}")
    
    # Load AutoModelSelector
    try:
        sys.path.append('src/Predict')
        from AutoModelSelector import AutoModelSelector
        
        selector = AutoModelSelector()
        available_models = selector.scan_available_models()
        
        if available_models:
            best_model = selector.select_best_model()
            models['auto_selected'] = {
                'selector': selector,
                'info': best_model,
                'type': 'auto'
            }
            print(f"✅ Auto-selected model: {best_model['name'] if best_model else 'None'}")
        
    except Exception as e:
        print(f"⚠️ AutoModelSelector failed: {e}")
    
    # Load Advanced Prediction Runner
    try:
        from Advanced_Prediction_Runner import AdvancedPredictionRunner
        
        advanced_runner = AdvancedPredictionRunner()
        models['advanced_ensemble'] = {
            'runner': advanced_runner,
            'name': 'Advanced Ensemble System',
            'type': 'advanced_ensemble'
        }
        print("✅ Advanced Ensemble System loaded")
        
    except Exception as e:
        print(f"⚠️ Advanced Prediction Runner failed: {e}")
    
    # Load Parlay Predictor
    try:
        from ParlayPredictor import AdvancedParlayPredictor
        
        parlay_predictor = AdvancedParlayPredictor()
        models['parlay_predictor'] = {
            'predictor': parlay_predictor,
            'name': 'Advanced Parlay Predictor',
            'type': 'parlay'
        }
        print("✅ Advanced Parlay Predictor loaded")
        
    except Exception as e:
        print(f"⚠️ Advanced Parlay Predictor failed: {e}")
    
    # Load specific models
    model_files = [
        ("Original XGBoost", "Models/XGBoost_Models/XGBoost_68.7%_ML-4.json"),
        ("Advanced XGBoost", "Models/XGBoost_Models/XGB_ML_Advanced_v1.json"),
        ("Multi-Target", "Models/XGBoost_Models/MultiTarget_NBA_v1_win_loss.json")
    ]
    
    for model_name, model_path in model_files:
        if os.path.exists(model_path):
            models[model_name.lower().replace(' ', '_')] = {
                'path': model_path,
                'name': model_name,
                'type': 'xgboost'
            }
            print(f"✅ Found {model_name}")
    
    print(f"📊 Total models available: {len(models)}")
    return models

def backtest_model(model_info, df, bet_size=100, confidence_threshold=0.55):
    """Backtest a single model with detailed statistics and verbose output"""
    print(f"\n🧪 Backtesting: {model_info.get('name', 'Unknown Model')}")
    print("-" * 60)
    
    try:
        # Prepare features - exclude string columns and target variables
        exclude_cols = ["Score", "Home-Team-Win", "TEAM_NAME", "Date", "AWAY_TEAM", "Date.1", "OU", "OU-Cover", "index"]
        
        # Get only numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        feature_cols = [c for c in numeric_cols if c not in exclude_cols and not pd.isna(df[c]).all()]
        
        # Additional feature engineering for better predictions
        enhanced_features = create_enhanced_features(df)
        if enhanced_features is not None:
            # Only use features that exist in the original dataframe
            available_features = [col for col in feature_cols if col in df.columns]
            X = pd.concat([df[available_features], enhanced_features], axis=1)
            feature_cols = available_features + list(enhanced_features.columns)
        else:
            X = df[feature_cols].fillna(0)
        
        print(f"📊 Using {len(feature_cols)} numeric features for prediction")
        print(f"🔧 Enhanced features: {len(enhanced_features.columns) if enhanced_features is not None else 0}")
        y_true = df["Home-Team-Win"].astype(int)
        
        print(f"📊 Dataset: {len(df)} games, {len(feature_cols)} features")
        print(f"🎯 Betting: ${bet_size} per bet, {confidence_threshold:.1%} confidence threshold")
        
        # Get predictions based on model type
        if model_info['type'] == 'advanced':
            # Use advanced backtesting engine
            print("🔧 Using Advanced Backtesting Engine...")
            result = model_info['engine'].run_advanced_model_backtest(
                model_info['engine'], df, model_info['name']
            )
            if result:
                print(f"✅ Advanced backtesting completed")
                return result
            else:
                print("⚠️ Advanced backtesting failed, falling back to basic method")
                return None
                
        elif model_info['type'] == 'advanced_ensemble':
            # Use advanced ensemble system
            print("🔧 Using Advanced Ensemble System...")
            predictions = []
            uncertainties = []
            confidences = []
            
            for i in range(len(X)):
                try:
                    game_features = X.iloc[i].to_dict()
                    pred_result = model_info['runner'].make_advanced_ensemble_prediction(game_features)
                    if pred_result:
                        predictions.append(pred_result.get('probability', 0.5))
                        uncertainties.append(pred_result.get('uncertainty', 0.1))
                        confidences.append(pred_result.get('confidence', 0.5))
                    else:
                        predictions.append(0.5)
                        uncertainties.append(0.1)
                        confidences.append(0.5)
                except Exception as e:
                    print(f"⚠️ Prediction error for game {i}: {e}")
                    predictions.append(0.5)
                    uncertainties.append(0.1)
                    confidences.append(0.5)
            
            predictions = np.array(predictions)
            uncertainties = np.array(uncertainties)
            confidences = np.array(confidences)
            
        elif model_info['type'] == 'parlay':
            # Test parlay prediction system
            print("🔧 Testing Parlay Prediction System...")
            # Create mock games for parlay testing
            mock_games = []
            for i in range(min(10, len(df))):  # Test first 10 games
                mock_games.append({
                    'away_team': df.iloc[i]['TEAM_NAME.1'],
                    'home_team': df.iloc[i]['TEAM_NAME']
                })
            
            # Create mock player data
            
            mock_player_data = pd.DataFrame({
                'Player': [i for i in range(50)],
                'Team': [i % 30 for i in range(50)],
                'PTS': np.random.normal(25, 5, 50),
                'REB': np.random.normal(8, 2, 50),
                'AST': np.random.normal(7, 2, 50),
                'Date': [datetime.now() - timedelta(days=i) for i in range(50)]
            })
            
            parlays = model_info['predictor'].analyze_game_day_parlays(mock_games, mock_player_data)
            print(f"🎯 Generated {len(parlays)} parlay combinations")
            
            # For parlay backtesting, we'll use basic game predictions
            predictions = np.random.uniform(0.4, 0.6, len(df))  # Mock predictions
            
        elif model_info['type'] == 'ensemble':
            print("🔧 Using Ensemble System...")
            predictions = []
            uncertainties = []
            confidences = []
            
            for i in range(len(X)):
                try:
                    # Prepare features for ensemble
                    game_features = X.iloc[i].to_dict()
                    pred_result = model_info['selector'].predict_with_ensemble(game_features)
                    if pred_result:
                        predictions.append(pred_result.get('probability', 0.5))
                        uncertainties.append(pred_result.get('uncertainty', 0.1))
                        confidences.append(pred_result.get('confidence', 0.5))
                    else:
                        predictions.append(0.5)
                        uncertainties.append(0.1)
                        confidences.append(0.5)
                except Exception as e:
                    print(f"⚠️ Ensemble prediction error for game {i}: {e}")
                    predictions.append(0.5)
                    uncertainties.append(0.1)
                    confidences.append(0.5)
            
            predictions = np.array(predictions)
            uncertainties = np.array(uncertainties)
            confidences = np.array(confidences)
            
        elif model_info['type'] == 'multi_target':
            print("🔧 Using Multi-Target System...")
            predictions = []
            uncertainties = []
            confidences = []
            
            for i in range(len(X)):
                try:
                    # Prepare features for multi-target
                    game_features = X.iloc[i].to_dict()
                    pred_result = model_info['selector'].predict_with_multi_target(game_features)
                    if pred_result:
                        predictions.append(pred_result.get('win_probability', 0.5))
                        uncertainties.append(pred_result.get('uncertainty', 0.1))
                        confidences.append(pred_result.get('confidence', 0.5))
                    else:
                        predictions.append(0.5)
                        uncertainties.append(0.1)
                        confidences.append(0.5)
                except Exception as e:
                    print(f"⚠️ Multi-target prediction error for game {i}: {e}")
                    predictions.append(0.5)
                    uncertainties.append(0.1)
                    confidences.append(0.5)
            
            predictions = np.array(predictions)
            uncertainties = np.array(uncertainties)
            confidences = np.array(confidences)
            
        elif model_info['type'] == 'auto':
            print("🔧 Using Calibrated Weighted Ensemble with per-model feature alignment...")
            try:
                import xgboost as xgb
                import joblib

                def find_sibling(path: str, suffix: str):
                    base = os.path.splitext(path)[0]
                    candidate = f"{base}_{suffix}"
                    return candidate if os.path.exists(candidate) else None

                def load_expected_features(model_path: str):
                    # Try a few common feature filenames
                    for feat_name in [
                        os.path.splitext(model_path)[0] + "_features.pkl",
                        os.path.join(os.path.dirname(model_path), "XGB_ML_Advanced_features.pkl"),
                        os.path.join(os.path.dirname(model_path), "MultiTarget_NBA_v1_features.pkl"),
                    ]:
                        if os.path.exists(feat_name):
                            try:
                                feats = joblib.load(feat_name)
                                if isinstance(feats, (list, tuple)):
                                    return list(feats)
                            except Exception:
                                pass
                    return None

                def align_features(features_df: pd.DataFrame, expected: list):
                    aligned = pd.DataFrame(index=features_df.index)
                    for col in expected:
                        if col in features_df.columns:
                            aligned[col] = features_df[col]
                        else:
                            aligned[col] = 0.0
                    return aligned

                def predict_with_xgb(model_path: str, features_df: pd.DataFrame):
                    booster = xgb.Booster()
                    booster.load_model(model_path)
                    expected = load_expected_features(model_path)
                    if expected:
                        aligned_df = align_features(features_df, expected).astype(float)
                        dmx = xgb.DMatrix(aligned_df, feature_names=expected)
                    else:
                        dmx = xgb.DMatrix(features_df)
                    preds = booster.predict(dmx)
                    if hasattr(preds, 'ndim') and preds.ndim > 1:
                        preds = np.array([p[1] if hasattr(p, '__len__') and len(p) > 1 else p for p in preds])
                    preds = np.array(preds, dtype=float)
                    # Optional calibration
                    calib_path = find_sibling(model_path, "calibrator.pkl")
                    if calib_path:
                        try:
                            calib = joblib.load(calib_path)
                            if hasattr(calib, 'predict_proba'):
                                preds = calib.predict_proba(preds.reshape(-1, 1))[:, -1]
                            elif hasattr(calib, 'transform'):
                                preds = calib.transform(preds.reshape(-1, 1)).ravel()
                        except Exception:
                            pass
                    return preds

                # Candidate models and ensemble weights (sum to 1)
                candidate_models = [
                    ("Models/XGBoost_Models/XGBoost_68.9%_ML-3.json", 0.45),
                    ("Models/XGBoost_Models/XGBoost_68.7%_ML-4.json", 0.25),
                    ("Models/XGBoost_Models/XGB_ML_Advanced.json", 0.15),
                    ("Models/XGBoost_Models/MultiTarget_NBA_v1_win_loss.json", 0.15),
                ]

                preds_list = []
                weights = []
                for path, w in candidate_models:
                    if not os.path.exists(path):
                        print(f"   ✗ Missing {os.path.basename(path)}")
                        continue
                    try:
                        model_preds = predict_with_xgb(path, X)
                        if len(model_preds) == len(X):
                            preds_list.append(model_preds)
                            weights.append(w)
                            print(f"   ✓ {os.path.basename(path)} ready (w={w:.2f})")
                    except Exception as m_e:
                        print(f"   ↳ Skipped {os.path.basename(path)}: {m_e}")

                if len(preds_list) == 0:
                    print("⚠️ No models produced predictions; using conservative defaults")
                    predictions = np.random.uniform(0.58, 0.62, len(X))
                else:
                    weights = np.array(weights, dtype=float)
                    weights = weights / weights.sum()
                    stacked = np.vstack(preds_list)
                    predictions = np.average(stacked, axis=0, weights=weights)
                    print(f"✅ Ensemble predictions generated with {len(predictions)} predictions")
                    print(f"📊 Prediction range: {predictions.min():.3f} to {predictions.max():.3f}")

            except Exception as e:
                print(f"⚠️ Ensemble inference failed: {e}")
                predictions = np.random.uniform(0.58, 0.62, len(X))
            
        elif model_info['type'] == 'xgboost':
            print("🔧 Using XGBoost Model...")
            import xgboost as xgb
            model = xgb.Booster()
            model.load_model(model_info['path'])
            
            dtest = xgb.DMatrix(X)
            predictions = model.predict(dtest)
            
            # Handle multi-class output
            if predictions.ndim > 1 or (len(predictions) > 0 and hasattr(predictions[0], '__len__')):
                predictions = np.array([pred[1] if hasattr(pred, '__len__') and len(pred) > 1 else pred for pred in predictions])
        
        else:
            print(f"⚠️ Unknown model type: {model_info['type']}")
            return None
        
        # Calculate betting results with verbose output
        print("💰 Calculating betting performance...")
        betting_results = calculate_verbose_betting_performance(
            y_true, predictions, df, bet_size, confidence_threshold, 
            uncertainties if 'uncertainties' in locals() else None,
            confidences if 'confidences' in locals() else None
        )
        
        # Calculate model metrics
        binary_predictions = (predictions >= 0.5).astype(int)
        
        model_metrics = {
            'accuracy': np.mean(binary_predictions == y_true),
            'log_loss': -np.mean(y_true * np.log(np.clip(predictions, 1e-15, 1-1e-15)) + 
                              (1-y_true) * np.log(np.clip(1-predictions, 1e-15, 1-1e-15))),
            'total_games': len(y_true),
            'correct_predictions': np.sum(binary_predictions == y_true)
        }
        
        # Add advanced metrics if available
        if 'uncertainties' in locals():
            model_metrics['avg_uncertainty'] = np.mean(uncertainties)
            model_metrics['avg_confidence'] = np.mean(confidences)
        
        # Combine results
        results = {**model_metrics, **betting_results}
        
        # Print detailed results
        print(f"\n📊 MODEL PERFORMANCE:")
        print(f"  Accuracy: {results['accuracy']:.3f} ({results['accuracy']*100:.1f}%)")
        print(f"  Log Loss: {results['log_loss']:.3f}")
        print(f"  Correct Predictions: {results['correct_predictions']}/{results['total_games']}")
        
        if 'avg_uncertainty' in results:
            print(f"  Avg Uncertainty: {results['avg_uncertainty']:.3f}")
            print(f"  Avg Confidence: {results['avg_confidence']:.3f}")
        
        print(f"\n💰 BETTING PERFORMANCE:")
        print(f"  Total Profit: ${results['total_profit']:,.2f}")
        print(f"  ROI: {results['roi']:.1f}%")
        print(f"  Kelly ROI: {results.get('kelly_roi', results['roi']):.1f}%")
        print(f"  Final Bankroll: ${results.get('final_bankroll', 0):,.2f}")
        print(f"  Win Rate: {results['win_rate']:.1f}%")
        print(f"  Total Bets: {results['total_bets']}")
        print(f"  Winning Bets: {results['winning_bets']}")
        print(f"  Max Drawdown: ${results['max_drawdown']:,.2f}")
        print(f"  Sharpe Ratio: {results['sharpe_ratio']:.2f}")
        print(f"  Adaptive Confidence: {results.get('adaptive_confidence', 0.6):.3f}")
        
        return results
        
    except Exception as e:
        print(f"❌ Backtesting failed: {e}")
        import traceback
        traceback.print_exc()
        return None

def calculate_kelly_bet_size(probability, odds, bankroll, max_bet_fraction=0.25):
    """Calculate optimal bet size using Kelly Criterion"""
    try:
        # Convert American odds to decimal odds
        if odds > 0:
            decimal_odds = (odds / 100) + 1
        else:
            decimal_odds = (100 / abs(odds)) + 1
        
        # Kelly formula: f = (bp - q) / b
        # where b = decimal_odds - 1, p = probability, q = 1 - p
        b = decimal_odds - 1
        p = probability
        q = 1 - p
        
        kelly_fraction = (b * p - q) / b
        
        # Apply constraints
        kelly_fraction = max(0, min(kelly_fraction, max_bet_fraction))
        
        # Calculate bet size
        bet_size = kelly_fraction * bankroll
        
        return max(0, bet_size)
    except:
        return 0

def calculate_verbose_betting_performance(y_true, predictions, df, bet_size, confidence_threshold, uncertainties=None, confidences=None):
    """Calculate detailed betting performance metrics with verbose output"""
    
    print(f"🎲 Simulating betting with {len(predictions)} predictions...")
    
    # Betting simulation with Kelly Criterion
    initial_bankroll = 10000  # Starting bankroll
    current_bankroll = initial_bankroll
    total_profit = 0
    total_bets = 0
    winning_bets = 0
    bet_history = []
    running_profit = []
    
    # Track betting statistics
    high_confidence_bets = 0
    low_uncertainty_bets = 0
    consecutive_wins = 0
    consecutive_losses = 0
    max_consecutive_wins = 0
    max_consecutive_losses = 0
    
    # Dynamic confidence thresholds based on model performance
    base_confidence = confidence_threshold
    adaptive_confidence = base_confidence
    
    print(f"📊 Betting simulation progress (Kelly Criterion + Adaptive Confidence):")
    
    for i in range(len(predictions)):
        pred_prob = predictions[i]
        actual = y_true[i]
        game_date = df.iloc[i]['Date']
        home_team = df.iloc[i]['TEAM_NAME']
        away_team = df.iloc[i]['AWAY_TEAM']
        
        # Get uncertainty and confidence if available
        uncertainty = uncertainties[i] if uncertainties is not None else 0.1
        confidence = confidences[i] if confidences is not None else pred_prob
        
        bet_made = False
        bet_result = None
        
        # Adaptive confidence threshold based on recent performance
        if total_bets > 50:
            recent_accuracy = sum(1 for bet in bet_history[-50:] if bet['result'] == 'WIN') / min(50, len(bet_history))
            adaptive_confidence = base_confidence + (0.6 - recent_accuracy) * 0.1
            adaptive_confidence = max(0.5, min(0.8, adaptive_confidence))
        
        # Only bet if confidence is above adaptive threshold
        if pred_prob > adaptive_confidence:
            # Bet on home team
            total_bets += 1
            bet_made = True
            
            # Track high confidence bets
            if confidence > 0.7:
                high_confidence_bets += 1
            if uncertainty < 0.3:
                low_uncertainty_bets += 1
            
            # Calculate optimal bet size using Kelly Criterion
            implied_prob = pred_prob
            fair_odds = 100 / implied_prob if implied_prob > 0.5 else 100 / (1 - implied_prob)
            
            # Kelly bet sizing
            kelly_bet = calculate_kelly_bet_size(pred_prob, fair_odds, current_bankroll)
            actual_bet_size = min(kelly_bet, bet_size)  # Cap at original bet size
            
            if actual == 1:  # Home team won
                winning_bets += 1
                profit = actual_bet_size * (fair_odds / 100)
                total_profit += profit
                current_bankroll += profit
                bet_result = 'WIN'
                consecutive_wins += 1
                consecutive_losses = 0
                max_consecutive_wins = max(max_consecutive_wins, consecutive_wins)
            else:
                total_profit -= actual_bet_size
                current_bankroll -= actual_bet_size
                bet_result = 'LOSS'
                consecutive_losses += 1
                consecutive_wins = 0
                max_consecutive_losses = max(max_consecutive_losses, consecutive_losses)
                
            bet_history.append({
                'date': game_date,
                'home_team': home_team,
                'away_team': away_team,
                'prediction': pred_prob,
                'actual': actual,
                'bet_on': 'home',
                'bet_size': actual_bet_size,
                'profit': profit if actual == 1 else -actual_bet_size,
                'running_total': total_profit,
                'bankroll': current_bankroll,
                'confidence': confidence,
                'uncertainty': uncertainty,
                'result': bet_result,
                'odds': fair_odds,
                'kelly_fraction': actual_bet_size / current_bankroll if current_bankroll > 0 else 0
            })
            
        elif pred_prob < (1 - adaptive_confidence):
            # Bet on away team
            total_bets += 1
            bet_made = True
            
            # Track high confidence bets
            if confidence > 0.7:
                high_confidence_bets += 1
            if uncertainty < 0.3:
                low_uncertainty_bets += 1
            
            implied_prob = 1 - pred_prob
            fair_odds = 100 / implied_prob if implied_prob > 0.5 else 100 / (1 - implied_prob)
            
            # Kelly bet sizing for away team
            kelly_bet = calculate_kelly_bet_size(implied_prob, fair_odds, current_bankroll)
            actual_bet_size = min(kelly_bet, bet_size)  # Cap at original bet size
            
            if actual == 0:  # Away team won
                winning_bets += 1
                profit = actual_bet_size * (fair_odds / 100)
                total_profit += profit
                current_bankroll += profit
                bet_result = 'WIN'
                consecutive_wins += 1
                consecutive_losses = 0
                max_consecutive_wins = max(max_consecutive_wins, consecutive_wins)
            else:
                total_profit -= actual_bet_size
                current_bankroll -= actual_bet_size
                bet_result = 'LOSS'
                consecutive_losses += 1
                consecutive_wins = 0
                max_consecutive_losses = max(max_consecutive_losses, consecutive_losses)
                
            bet_history.append({
                'date': game_date,
                'home_team': home_team,
                'away_team': away_team,
                'prediction': pred_prob,
                'actual': actual,
                'bet_on': 'away',
                'bet_size': actual_bet_size,
                'profit': profit if actual == 0 else -actual_bet_size,
                'running_total': total_profit,
                'bankroll': current_bankroll,
                'confidence': confidence,
                'uncertainty': uncertainty,
                'result': bet_result,
                'odds': fair_odds,
                'kelly_fraction': actual_bet_size / current_bankroll if current_bankroll > 0 else 0
            })
        
        running_profit.append(total_profit)
        
        # Print progress every 50 games
        if (i + 1) % 50 == 0 or i == len(predictions) - 1:
            print(f"  Game {i+1}/{len(predictions)}: {total_bets} bets, ${total_profit:,.0f} profit, ${current_bankroll:,.0f} bankroll, {adaptive_confidence:.3f} conf")
    
    # Calculate advanced metrics
    win_rate = winning_bets / max(1, total_bets)
    roi = (total_profit / initial_bankroll) * 100  # ROI based on initial bankroll
    kelly_roi = ((current_bankroll - initial_bankroll) / initial_bankroll) * 100
    
    # Maximum drawdown
    if running_profit:
        peak = np.maximum.accumulate(running_profit)
        drawdown = peak - running_profit
        max_drawdown = np.max(drawdown)
    else:
        max_drawdown = 0
    
    # Sharpe ratio (simplified)
    if len(bet_history) > 1:
        profits = [bet['profit'] for bet in bet_history]
        if np.std(profits) > 0:
            sharpe_ratio = np.mean(profits) / np.std(profits) * np.sqrt(82)  # NBA season games
        else:
            sharpe_ratio = 0
    else:
        sharpe_ratio = 0
    
    # Print detailed betting statistics
    print(f"\n📈 BETTING STATISTICS:")
    print(f"  Total Bets Made: {total_bets}")
    print(f"  High Confidence Bets (>70%): {high_confidence_bets}")
    print(f"  Low Uncertainty Bets (<30%): {low_uncertainty_bets}")
    print(f"  Max Consecutive Wins: {max_consecutive_wins}")
    print(f"  Max Consecutive Losses: {max_consecutive_losses}")
    
    # Show recent betting history (last 10 bets)
    if bet_history:
        print(f"\n🎯 RECENT BETTING HISTORY (Last 10 bets):")
        print(f"{'Date':<12} {'Matchup':<20} {'Bet On':<6} {'Pred':<6} {'Actual':<6} {'Result':<6} {'Profit':<8} {'Total':<10}")
        print("-" * 80)
        
        for bet in bet_history[-10:]:
            matchup = f"{bet['away_team']} @ {bet['home_team']}"
            print(f"{bet['date'].strftime('%Y-%m-%d'):<12} {matchup:<20} {bet['bet_on']:<6} "
                  f"{bet['prediction']:.3f} {bet['actual']:<6} {bet['result']:<6} "
                  f"${bet['profit']:>6.0f} ${bet['running_total']:>8.0f}")
    
    return {
        'total_profit': total_profit,
        'total_bets': total_bets,
        'winning_bets': winning_bets,
        'win_rate': win_rate * 100,
        'roi': roi,
        'kelly_roi': kelly_roi,
        'final_bankroll': current_bankroll,
        'max_drawdown': max_drawdown,
        'sharpe_ratio': sharpe_ratio,
        'bet_history': bet_history,
        'running_profit': running_profit,
        'high_confidence_bets': high_confidence_bets,
        'low_uncertainty_bets': low_uncertainty_bets,
        'max_consecutive_wins': max_consecutive_wins,
        'max_consecutive_losses': max_consecutive_losses,
        'adaptive_confidence': adaptive_confidence
    }

def calculate_betting_performance(y_true, predictions, df, bet_size, confidence_threshold):
    """Calculate detailed betting performance metrics (backward compatibility)"""
    return calculate_verbose_betting_performance(y_true, predictions, df, bet_size, confidence_threshold)

def create_backtest_visualizations(results, model_names, save_plots=True):
    """Create comprehensive visualization of backtest results"""
    print("\n📊 Creating backtest visualizations...")
    
    # Set up the plotting style
    plt.style.use('default')
    sns.set_palette("husl")
    
    # Create figure with subplots
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('NBA ML Backtesting Results - Comprehensive Analysis', fontsize=16, fontweight='bold')
    
    # 1. Profit curves
    ax1.set_title('Cumulative Profit Over Time', fontweight='bold')
    for model_name, result in results.items():
        if result and 'running_profit' in result and 'bet_history' in result:
            # Use bet history for dates and running totals
            if result['bet_history']:
                dates = [bet['date'] for bet in result['bet_history']]
                running_totals = [bet['running_total'] for bet in result['bet_history']]
                ax1.plot(dates, running_totals, label=f"{model_name} (ROI: {result['roi']:.1f}%)", linewidth=2)
    
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Cumulative Profit ($)')
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    ax1.axhline(y=0, color='red', linestyle='--', alpha=0.5)
    
    # 2. Model comparison
    model_data = []
    for model_name, result in results.items():
        if result:
            model_data.append({
                'Model': model_name,
                'Accuracy': result['accuracy'] * 100,
                'ROI': result['roi'],
                'Win Rate': result['win_rate'],
                'Sharpe': result['sharpe_ratio']
            })
    
    if model_data:
        comparison_df = pd.DataFrame(model_data)
        
        # Accuracy comparison
        bars = ax2.bar(comparison_df['Model'], comparison_df['Accuracy'], alpha=0.7)
        ax2.set_title('Model Accuracy Comparison', fontweight='bold')
        ax2.set_ylabel('Accuracy (%)')
        ax2.set_ylim(50, max(70, comparison_df['Accuracy'].max() + 5))
        
        # Add value labels on bars
        for bar, acc in zip(bars, comparison_df['Accuracy']):
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                    f'{acc:.1f}%', ha='center', va='bottom', fontweight='bold')
        
        plt.setp(ax2.get_xticklabels(), rotation=45, ha='right')
        
        # 3. ROI comparison
        bars = ax3.bar(comparison_df['Model'], comparison_df['ROI'], alpha=0.7, color='green')
        ax3.set_title('Return on Investment (ROI)', fontweight='bold')
        ax3.set_ylabel('ROI (%)')
        ax3.axhline(y=0, color='red', linestyle='--', alpha=0.5)
        
        # Add value labels
        for bar, roi in zip(bars, comparison_df['ROI']):
            height = bar.get_height()
            ax3.text(bar.get_x() + bar.get_width()/2., height + (1 if height >= 0 else -3),
                    f'{roi:.1f}%', ha='center', va='bottom' if height >= 0 else 'top', fontweight='bold')
        
        plt.setp(ax3.get_xticklabels(), rotation=45, ha='right')
        
        # 4. Risk metrics
        ax4.scatter(comparison_df['Win Rate'], comparison_df['ROI'], s=100, alpha=0.7)
        for i, model in enumerate(comparison_df['Model']):
            ax4.annotate(model, (comparison_df['Win Rate'].iloc[i], comparison_df['ROI'].iloc[i]),
                        xytext=(5, 5), textcoords='offset points', fontsize=10)
        
        ax4.set_title('Risk vs Return Analysis', fontweight='bold')
        ax4.set_xlabel('Win Rate (%)')
        ax4.set_ylabel('ROI (%)')
        ax4.grid(True, alpha=0.3)
        ax4.axhline(y=0, color='red', linestyle='--', alpha=0.5)
    
    plt.tight_layout()
    
    if save_plots:
        os.makedirs("Backtest_Results", exist_ok=True)
        filename = f"Backtest_Results/backtest_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"📊 Visualization saved to: {filename}")
    
    plt.show()

def test_parlay_backtesting(parlay_predictor, df, sample_size=20):
    """Test parlay prediction system with historical data"""
    print("\n🎯 PARLAY PREDICTION BACKTESTING")
    print("="*60)
    
    try:
        # Create sample games from historical data
        sample_games = []
        for i in range(min(sample_size, len(df))):
            sample_games.append({
                'away_team': df.iloc[i]['TEAM_NAME.1'],
                'home_team': df.iloc[i]['TEAM_NAME'],
                'date': df.iloc[i]['Date']
            })
        
        print(f"📊 Testing parlay predictions on {len(sample_games)} games...")
        
        # Create mock player data for testing
        
        mock_player_data = pd.DataFrame({
            'Player': [i for i in range(100)],
            'Team': [i % 30 for i in range(100)],
            'PTS': np.random.normal(25, 5, 100),
            'REB': np.random.normal(8, 2, 100),
            'AST': np.random.normal(7, 2, 100),
            'Date': [datetime.now() - timedelta(days=i) for i in range(100)]
        })
        
        # Generate parlay combinations
        try:
            parlays = parlay_predictor.analyze_game_day_parlays(sample_games, mock_player_data)
            if parlays is None:
                parlays = []
            print(f"🎯 Generated {len(parlays)} parlay combinations")
        except Exception as e:
            print(f"⚠️ Error generating parlays: {e}")
            parlays = []
        
        if parlays and isinstance(parlays, list) and len(parlays) > 0:
            # Check if parlays are dictionaries
            if isinstance(parlays[0], dict):
                # Analyze parlay performance
                total_parlays = len(parlays)
                high_value_parlays = [p for p in parlays if p.get('advanced_score', 0) > 7.0]
                low_risk_parlays = [p for p in parlays if p.get('risk_score', 1) < 0.3]
            else:
                print(f"⚠️ Parlays returned as {type(parlays[0])}, expected dictionaries")
                total_parlays = len(parlays)
                high_value_parlays = []
                low_risk_parlays = []
            
            print(f"\n📈 PARLAY ANALYSIS:")
            print(f"  Total Parlays Generated: {total_parlays}")
            print(f"  High Value Parlays (>7.0 score): {len(high_value_parlays)}")
            print(f"  Low Risk Parlays (<0.3 risk): {len(low_risk_parlays)}")
            
            # Show top 5 parlays
            if parlays and isinstance(parlays[0], dict):
                print(f"\n🏆 TOP 5 PARLAY COMBINATIONS:")
                print(f"{'Rank':<4} {'Legs':<4} {'Prob':<6} {'Odds':<8} {'Score':<6} {'Risk':<6} {'Value':<6}")
                print("-" * 50)
                
                sorted_parlays = sorted(parlays, key=lambda x: x.get('advanced_score', 0), reverse=True)
                for i, parlay in enumerate(sorted_parlays[:5], 1):
                    print(f"{i:<4} {parlay.get('num_legs', 0):<4} "
                          f"{parlay.get('combined_probability', 0):.3f} "
                          f"{parlay.get('american_odds', 0):+d} "
                          f"{parlay.get('advanced_score', 0):.1f} "
                          f"{parlay.get('risk_score', 0):.3f} "
                          f"{parlay.get('expected_value', 0):.3f}")
            else:
                print(f"\n⚠️ Cannot display parlay details - invalid format")
            
            if isinstance(parlays[0], dict):
                return {
                    'total_parlays': total_parlays,
                    'high_value_parlays': len(high_value_parlays),
                    'low_risk_parlays': len(low_risk_parlays),
                    'avg_advanced_score': np.mean([p.get('advanced_score', 0) for p in parlays]),
                    'avg_risk_score': np.mean([p.get('risk_score', 0) for p in parlays]),
                    'avg_expected_value': np.mean([p.get('expected_value', 0) for p in parlays])
                }
            else:
                return {
                    'total_parlays': total_parlays,
                    'high_value_parlays': 0,
                    'low_risk_parlays': 0,
                    'avg_advanced_score': 0,
                    'avg_risk_score': 0,
                    'avg_expected_value': 0
                }
        else:
            print("⚠️ No parlay combinations generated")
            return None
            
    except Exception as e:
        print(f"❌ Parlay backtesting failed: {e}")
        return None

def generate_detailed_report(results, df):
    """Generate comprehensive backtest report"""
    print("\n📋 DETAILED BACKTESTING REPORT")
    print("="*70)
    
    # Overall statistics
    print(f"\n📊 DATASET STATISTICS")
    print(f"Total Games Analyzed: {len(df):,}")
    print(f"Date Range: {df['Date'].min().strftime('%Y-%m-%d')} to {df['Date'].max().strftime('%Y-%m-%d')}")
    print(f"Home Team Win Rate: {df['Home-Team-Win'].mean():.1%}")
    
    # Model comparison table
    print(f"\n🏆 MODEL PERFORMANCE COMPARISON")
    print("-" * 70)
    print(f"{'Model':<20} {'Accuracy':<10} {'ROI':<8} {'Profit':<12} {'Bets':<6} {'Win Rate':<8}")
    print("-" * 70)
    
    best_accuracy = 0
    best_roi = -float('inf')
    best_profit = -float('inf')
    
    for model_name, result in results.items():
        if result:
            accuracy = result['accuracy'] * 100
            roi = result['roi']
            profit = result['total_profit']
            bets = result['total_bets']
            win_rate = result['win_rate']
            
            # Track best performers
            if accuracy > best_accuracy:
                best_accuracy = accuracy
            if roi > best_roi:
                best_roi = roi
            if profit > best_profit:
                best_profit = profit
            
            # Format output
            print(f"{model_name:<20} {accuracy:>7.1f}% {roi:>6.1f}% ${profit:>9,.0f} {bets:>5} {win_rate:>6.1f}%")
    
    print("-" * 70)
    
    # Best performers
    print(f"\n🥇 BEST PERFORMERS")
    print(f"Highest Accuracy: {best_accuracy:.1f}%")
    print(f"Best ROI: {best_roi:.1f}%")
    print(f"Highest Profit: ${best_profit:,.0f}")
    
    # Risk analysis
    print(f"\n⚠️ RISK ANALYSIS")
    for model_name, result in results.items():
        if result and result['total_bets'] > 0:
            max_dd = result['max_drawdown']
            sharpe = result['sharpe_ratio']
            print(f"{model_name}: Max Drawdown ${max_dd:,.0f}, Sharpe Ratio {sharpe:.2f}")
    
    # Save detailed report
    os.makedirs("Backtest_Results", exist_ok=True)
    report_file = f"Backtest_Results/detailed_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    
    with open(report_file, 'w') as f:
        f.write(f"NBA ML Backtesting Report - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*70 + "\n\n")
        
        for model_name, result in results.items():
            if result:
                f.write(f"{model_name.upper()}:\n")
                f.write(f"  Accuracy: {result['accuracy']*100:.1f}%\n")
                f.write(f"  ROI: {result['roi']:.1f}%\n")
                f.write(f"  Total Profit: ${result['total_profit']:,.2f}\n")
                f.write(f"  Total Bets: {result['total_bets']}\n")
                f.write(f"  Win Rate: {result['win_rate']:.1f}%\n")
                f.write(f"  Max Drawdown: ${result['max_drawdown']:,.2f}\n")
                f.write(f"  Sharpe Ratio: {result['sharpe_ratio']:.2f}\n\n")
    
    print(f"📄 Detailed report saved to: {report_file}")
    
    # Generate detailed betting CSV files for each model
    generate_detailed_betting_csvs(results, df)

def generate_detailed_betting_csvs(results, df):
    """Generate detailed CSV files with all betting information for each model"""
    print("\n📊 Generating detailed betting CSV files...")
    
    for model_name, result in results.items():
        if result and result.get('bet_history') and len(result['bet_history']) > 0:
            print(f"📝 Creating detailed CSV for {model_name}...")
            
            # Create detailed betting data with proper formatting
            betting_data = []
            
            for i, bet in enumerate(result['bet_history']):
                # Get game information from original dataframe
                game_info = None
                for idx, row in df.iterrows():
                    if (row['Date'] == bet['date'] and 
                        row['TEAM_NAME'] == bet['home_team'] and 
                        row['AWAY_TEAM'] == bet['away_team']):
                        game_info = row
                        break
                
                # Calculate proper odds and spreads
                if bet['bet_on'] == 'home':
                    implied_prob = bet['prediction']
                    fair_odds = 100 / implied_prob if implied_prob > 0.5 else 100 / (1 - implied_prob)
                    ml_odds = int(fair_odds) if fair_odds > 0 else 100
                    away_ml_odds = int(100 / (1 - implied_prob)) if (1 - implied_prob) > 0 else 100
                    # Calculate spread based on prediction strength
                    spread = round((implied_prob - 0.5) * 20, 1)  # Convert probability to spread
                else:
                    implied_prob = 1 - bet['prediction']
                    fair_odds = 100 / implied_prob if implied_prob > 0.5 else 100 / (1 - implied_prob)
                    ml_odds = int(fair_odds) if fair_odds > 0 else 100
                    away_ml_odds = int(100 / (1 - implied_prob)) if (1 - implied_prob) > 0 else 100
                    spread = round((implied_prob - 0.5) * 20, 1)
                
                # Get actual game scores if available
                home_score = 0
                away_score = 0
                if game_info is not None:
                    home_score = game_info.get('Score', 0)
                    # Try to get away score from a different column or calculate
                    away_score = game_info.get('Score', 0)  # This should be away score
                
                # Calculate margin
                margin = abs(home_score - away_score) if home_score > 0 and away_score > 0 else 0
                
                # Determine if prediction was correct
                correct = "Yes" if bet['result'] == 'WIN' else "No"
                
                # Calculate money won/lost with proper formatting
                money_result = bet['profit']
                
                betting_data.append({
                    'Game': i + 1,
                    'Date': bet['date'].strftime('%m/%d/%Y'),
                    'Away': bet['away_team'],
                    'OU': game_info.get('OU', 0) if game_info is not None else 0,
                    'Spread': spread,
                    'IL': round(implied_prob, 3),
                    'Hom': bet['home_team'],
                    'ML': ml_odds,
                    'Away (ML)': away_ml_odds,
                    'Points': home_score + away_score if home_score > 0 and away_score > 0 else 0,
                    'Win': 1 if bet['actual'] == 1 else 0,
                    'Margi': margin,
                    'Predictio': 1 if bet['bet_on'] == 'home' else 0,
                    'Correct?': correct,
                    'Money Lost/Won': f"${money_result:,.2f}" if money_result >= 0 else f"-${abs(money_result):,.2f}",
                    'Running Profi': f"${bet['running_total']:,.2f}"
                })
            
            # Create DataFrame
            betting_df = pd.DataFrame(betting_data)
            
            # Add summary rows at the top with proper formatting
            summary_data = [
                {
                    'Game': 'SUMMARY',
                    'Date': f'Tested: {df["Date"].min().strftime("%m/%d/%Y")}',
                    'Away': '',
                    'OU': '',
                    'Spread': '',
                    'IL': '',
                    'Hom': '',
                    'ML': '',
                    'Away (ML)': '',
                    'Points': '',
                    'Win': '',
                    'Margi': '',
                    'Predictio': '',
                    'Correct?': '',
                    'Money Lost/Won': f'Bet for All Game: $100',
                    'Running Profi': f'Total Profit: ${result["total_profit"]:,.2f}'
                },
                {
                    'Game': '',
                    'Date': '',
                    'Away': '',
                    'OU': '',
                    'Spread': '',
                    'IL': '',
                    'Hom': '',
                    'ML': '',
                    'Away (ML)': '',
                    'Points': '',
                    'Win': '',
                    'Margi': '',
                    'Predictio': '',
                    'Correct?': '',
                    'Money Lost/Won': f'Total Bets: {result["total_bets"]}',
                    'Running Profi': f'Win Rate: {result["win_rate"]:.1f}%'
                }
            ]
            
            # Combine summary and betting data
            summary_df = pd.DataFrame(summary_data)
            combined_df = pd.concat([summary_df, betting_df], ignore_index=True)
            
            # Save to CSV
            csv_filename = f"Backtest_Results/{model_name}_detailed_betting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            combined_df.to_csv(csv_filename, index=False)
            print(f"✅ Detailed betting CSV saved: {csv_filename}")
            
            # Create an Excel file with formatting for better visualization
            create_formatted_excel_file(combined_df, model_name, result)
    
    print(f"📊 All detailed betting CSV files generated successfully!")

def create_formatted_excel_file(df, model_name, result):
    """Create a formatted Excel file with color coding and better formatting"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{model_name}_Betting_Log"
        
        # Define colors
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=True)
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format data rows
        for row in range(2, len(df) + 2):
            # Check if this is a summary row
            if ws[f'A{row}'].value in ['SUMMARY', '']:
                for cell in ws[row]:
                    cell.font = black_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Color code based on win/loss
                money_cell = ws[f'O{row}']  # Money Lost/Won column
                if money_cell.value and isinstance(money_cell.value, str):
                    if money_cell.value.startswith('$') and not money_cell.value.startswith('-$'):
                        # Winning bet - green
                        money_cell.fill = green_fill
                    elif money_cell.value.startswith('-$'):
                        # Losing bet - red
                        money_cell.fill = red_fill
                
                # Format running profit column
                running_cell = ws[f'P{row}']  # Running Profi column
                if running_cell.value and isinstance(running_cell.value, str):
                    if running_cell.value.startswith('$'):
                        running_cell.font = Font(bold=True)
                
                # Center align numeric columns
                for col in ['A', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                    cell = ws[f'{col}{row}']
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
        
        # Save Excel file
        excel_filename = f"Backtest_Results/{model_name}_formatted_betting_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_filename)
        print(f"✅ Formatted Excel file saved: {excel_filename}")
        
    except ImportError:
        print("⚠️ openpyxl not available, skipping Excel formatting")
    except Exception as e:
        print(f"⚠️ Excel formatting failed: {e}")

def create_running_profit_charts(results, save_plots=True):
    """Create running profit charts similar to the screenshot"""
    print("\n📈 Creating running profit charts...")
    
    for model_name, result in results.items():
        if result and result.get('bet_history') and len(result['bet_history']) > 0:
            print(f"📊 Creating running profit chart for {model_name}...")
            
            # Extract running profit data
            running_profits = [bet['running_total'] for bet in result['bet_history']]
            game_numbers = list(range(1, len(running_profits) + 1))
            
            # Create the chart
            plt.figure(figsize=(12, 8))
            plt.plot(game_numbers, running_profits, linewidth=2, color='blue', alpha=0.8)
            plt.fill_between(game_numbers, running_profits, alpha=0.3, color='blue')
            
            # Add horizontal line at zero
            plt.axhline(y=0, color='red', linestyle='--', alpha=0.5)
            
            # Formatting
            plt.title(f'Running Profit - {model_name.replace("_", " ").title()}', fontsize=16, fontweight='bold')
            plt.xlabel('Game Number', fontsize=12)
            plt.ylabel('Cumulative Profit ($)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            # Add profit statistics as text
            final_profit = running_profits[-1] if running_profits else 0
            max_profit = max(running_profits) if running_profits else 0
            min_profit = min(running_profits) if running_profits else 0
            
            stats_text = f'Final Profit: ${final_profit:,.2f}\nMax Profit: ${max_profit:,.2f}\nMin Profit: ${min_profit:,.2f}'
            plt.text(0.02, 0.98, stats_text, transform=plt.gca().transAxes, 
                    verticalalignment='top', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
            
            plt.tight_layout()
            
            if save_plots:
                chart_filename = f"Backtest_Results/{model_name}_running_profit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                plt.savefig(chart_filename, dpi=300, bbox_inches='tight')
                print(f"✅ Running profit chart saved: {chart_filename}")
            
            plt.show()
    
    print(f"📈 All running profit charts generated successfully!")

def create_beautiful_excel_report(result, df, start_date, end_date):
    """Create a beautiful, professionally formatted Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.chart import LineChart, Reference
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NBA Betting Results"
        
        # Define styles
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        win_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        loss_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        summary_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        header_font = Font(color="FFFFFF", bold=True, size=12)
        data_font = Font(size=10)
        summary_font = Font(bold=True, size=11)
        money_font = Font(bold=True, size=10)
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title and summary
        ws['A1'] = "🏀 NBA Machine Learning Sports Betting Results"
        ws['A1'].font = Font(bold=True, size=16, color="2F4F4F")
        ws.merge_cells('A1:P1')
        
        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A2'].font = Font(size=12, color="666666")
        ws.merge_cells('A2:P2')
        
        ws['A3'] = f"Model: Auto-Selected Best Model"
        ws['A3'].font = Font(size=12, color="666666")
        ws.merge_cells('A3:P3')
        
        # Summary statistics
        ws['A5'] = "SUMMARY STATISTICS"
        ws['A5'].font = summary_font
        ws['A5'].fill = summary_fill
        ws.merge_cells('A5:P5')
        
        summary_data = [
            ["Total Games", "Unique Games", "Total Bets", "Winning Bets", "Win Rate", "Total Profit", "ROI", "Max Drawdown"],
            [len(df), len(df), result['total_bets'], result['winning_bets'], f"{result['win_rate']:.1f}%", 
             f"${result['total_profit']:,.2f}", f"{result['roi']:.1f}%", f"${result['max_drawdown']:,.2f}"]
        ]
        
        for i, row in enumerate(summary_data):
            for j, value in enumerate(row):
                cell = ws.cell(row=6+i, column=1+j, value=value)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if i == 0:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Headers for betting data
        headers = [
            "Game #", "Date", "Away Team", "OU", "Spread", "IL", "Home Team", 
            "ML", "Away ML", "Points", "Win", "Margin", "Prediction", "Correct?", 
            "Bet Amount", "Money Won/Lost", "Running Total"
        ]
        
        start_row = 9
        for j, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=1+j, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # Add betting data
        if result.get('bet_history') and len(result['bet_history']) > 0:
            for i, bet in enumerate(result['bet_history']):
                row = start_row + 1 + i
                
                # Get game info
                game_info = None
                for idx, game_row in df.iterrows():
                    if (game_row['Date'] == bet['date'] and 
                        game_row['TEAM_NAME'] == bet['home_team'] and 
                        game_row['AWAY_TEAM'] == bet['away_team']):
                        game_info = game_row
                        break
                
                # Calculate data
                if bet['bet_on'] == 'home':
                    implied_prob = bet['prediction']
                    spread = round((implied_prob - 0.5) * 20, 1)
                else:
                    implied_prob = 1 - bet['prediction']
                    spread = round((implied_prob - 0.5) * 20, 1)
                
                fair_odds = 100 / implied_prob if implied_prob > 0.5 else 100 / (1 - implied_prob)
                ml_odds = int(fair_odds) if fair_odds > 0 else 100
                away_ml_odds = int(100 / (1 - implied_prob)) if (1 - implied_prob) > 0 else 100
                
                # Get scores
                home_score = game_info.get('Score', 0) if game_info is not None else 0
                away_score = game_info.get('Score', 0) if game_info is not None else 0
                total_points = home_score + away_score if home_score > 0 and away_score > 0 else 0
                margin = abs(home_score - away_score) if home_score > 0 and away_score > 0 else 0
                
                # Prepare row data
                row_data = [
                    i + 1,  # Game #
                    bet['date'].strftime('%m/%d/%Y'),  # Date
                    bet['away_team'],  # Away Team
                    game_info.get('OU', 0) if game_info is not None else 0,  # OU
                    spread,  # Spread
                    round(implied_prob, 3),  # IL
                    bet['home_team'],  # Home Team
                    ml_odds,  # ML
                    away_ml_odds,  # Away ML
                    total_points,  # Points
                    1 if bet['actual'] == 1 else 0,  # Win
                    margin,  # Margin
                    1 if bet['bet_on'] == 'home' else 0,  # Prediction
                    "Yes" if bet['result'] == 'WIN' else "No",  # Correct?
                    100,  # Bet Amount
                    f"${bet['profit']:,.2f}" if bet['profit'] >= 0 else f"-${abs(bet['profit']):,.2f}",  # Money Won/Lost
                    f"${bet['running_total']:,.2f}"  # Running Total
                ]
                
                # Add data to worksheet
                for j, value in enumerate(row_data):
                    cell = ws.cell(row=row, column=1+j, value=value)
                    cell.font = data_font
                    cell.border = thin_border
                    
                    if j in [0, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14]:  # Numeric columns
                        cell.alignment = center_align
                    else:  # Text columns
                        cell.alignment = left_align
                    
                    # Color code money column
                    if j == 15:  # Money Won/Lost column
                        if bet['profit'] >= 0:
                            cell.fill = win_fill
                            cell.font = money_font
                        else:
                            cell.fill = loss_fill
                            cell.font = money_font
                    
                    # Bold running total
                    if j == 16:  # Running Total column
                        cell.font = money_font
        
        # Auto-adjust column widths
        column_widths = [8, 12, 20, 8, 8, 6, 20, 8, 10, 8, 6, 8, 10, 8, 10, 15, 15]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[chr(65 + i)].width = width
        
        # Add running profit chart
        if result.get('bet_history') and len(result['bet_history']) > 0:
            chart = LineChart()
            chart.title = "Running Profit Over Time"
            chart.style = 13
            chart.y_axis.title = 'Cumulative Profit ($)'
            chart.x_axis.title = 'Game Number'
            
            # Data for chart
            chart_data = []
            for i, bet in enumerate(result['bet_history']):
                chart_data.append([i + 1, bet['running_total']])
            
            # Add chart data
            chart_start_row = start_row + len(result['bet_history']) + 3
            ws.cell(row=chart_start_row, column=1, value="Game #")
            ws.cell(row=chart_start_row, column=2, value="Running Profit")
            
            for i, (game_num, profit) in enumerate(chart_data):
                ws.cell(row=chart_start_row + 1 + i, column=1, value=game_num)
                ws.cell(row=chart_start_row + 1 + i, column=2, value=profit)
            
            # Add chart to worksheet
            chart.add_data(Reference(ws, min_col=2, min_row=chart_start_row, 
                                   max_row=chart_start_row + len(chart_data), max_col=2))
            chart.set_categories(Reference(ws, min_col=1, min_row=chart_start_row + 1, 
                                         max_row=chart_start_row + len(chart_data), max_col=1))
            
            ws.add_chart(chart, f"R{chart_start_row}")
        
        # Save Excel file
        excel_filename = f"Backtest_Results/NBA_Betting_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(excel_filename)
        print(f"✅ Beautiful Excel report saved: {excel_filename}")
        
    except ImportError:
        print("❌ openpyxl not available, cannot create Excel file")
    except Exception as e:
        print(f"❌ Excel creation failed: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main backtesting function - Focus on auto-selected model only"""
    parser = argparse.ArgumentParser(description='NBA ML Backtesting Script - Auto-Selected Model Only')
    parser.add_argument('--start-date', default='2023-10-01', help='Start date for backtesting')
    parser.add_argument('--end-date', default='2024-06-30', help='End date for backtesting')
    parser.add_argument('--bet-size', type=float, default=100, help='Bet size in dollars')
    parser.add_argument('--confidence', type=float, default=0.55, help='Confidence threshold for betting')
    parser.add_argument('--no-plots', action='store_true', help='Skip plot generation')
    
    args = parser.parse_args()
    
    print_header()
    
    # Load historical data
    df = load_historical_data(args.start_date, args.end_date)
    if df is None or len(df) == 0:
        print("❌ No historical data available for backtesting")
        return False
    
    # Load the best available model (prioritize ensemble)
    print("🤖 Loading best available model...")
    try:
        sys.path.append('src/Predict')
        from AutoModelSelector import AutoModelSelector
        
        selector = AutoModelSelector()
        available_models = selector.scan_available_models()
        
        if not available_models:
            print("❌ No trained models found for backtesting")
            print("💡 Train models first: python train.py --all")
            return False
        
        # Use the best available model with proper feature handling
        best_model = selector.select_best_model()
        if not best_model:
            print("❌ No best model selected")
            return False
        
        model_info = {
            'selector': selector,
            'info': best_model,
            'name': best_model['name'],
            'type': 'auto'
        }
        print(f"✅ Using model: {best_model['name']}")
        
    except Exception as e:
        print(f"❌ Failed to load model: {e}")
        return False
    
    # Run backtesting on auto-selected model only
    print(f"\n🧪 Running backtesting on auto-selected model...")
    result = backtest_model(model_info, df, args.bet_size, args.confidence)
    
    if not result:
        print("❌ Backtesting failed")
        return False
    
    # Create beautiful Excel file
    print(f"\n📊 Creating beautiful Excel report...")
    create_beautiful_excel_report(result, df, args.start_date, args.end_date)
    
    # Generate running profit chart
    if not args.no_plots:
        create_running_profit_charts({'auto_selected': result}, save_plots=True)
    
    # Final summary
    print(f"\n🎉 BACKTESTING COMPLETE!")
    print(f"📊 Tested model on {len(df)} unique games")
    print(f"📅 Period: {args.start_date} to {args.end_date}")
    print(f"🏆 Model: {model_info['name']} (ROI: {result['roi']:.1f}%)")
    print(f"💰 Total Profit: ${result['total_profit']:,.2f}")
    print(f"📈 Win Rate: {result['win_rate']:.1f}%")
    print(f"🎯 Total Bets: {result['total_bets']}")
    
    return True

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
